from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, text
from uuid import UUID
from decimal import Decimal
from pydantic import BaseModel
from typing import Optional
from backend.app.core.database import get_db
from backend.app.api.v1.routers.admin_auth import verify_admin_token
from backend.app.models.unit import Unit
from backend.app.models.project import Project
from backend.app.models.tower import Tower
import json
import csv
import io

router = APIRouter(prefix="/admin", tags=["admin"])

@router.get("/units")
async def list_units(
    page: int = 1, page_size: int = 20,
    status: str = "", unit_type: str = "",
    db: AsyncSession = Depends(get_db), admin=Depends(verify_admin_token)
):
    q = select(Unit)
    if status:    q = q.where(Unit.status == status)
    if unit_type: q = q.where(Unit.unit_type == unit_type)
    total = (await db.execute(select(func.count()).select_from(q.subquery()))).scalar()
    result = await db.execute(
        q.order_by(Unit.created_at.desc())
        .offset((page-1)*page_size).limit(page_size)
    )
    units = result.scalars().all()

    # Resolve tower + project names for display
    tower_ids = {u.tower_id for u in units if u.tower_id}
    tower_map: dict = {}
    project_map: dict = {}
    if tower_ids:
        trows = (await db.execute(select(Tower).where(Tower.id.in_(tower_ids)))).scalars().all()
        tower_map = {t.id: t for t in trows}
        project_ids = {t.project_id for t in trows if t.project_id}
        if project_ids:
            prows = (await db.execute(select(Project).where(Project.id.in_(project_ids)))).scalars().all()
            project_map = {p.id: p for p in prows}

    def _row(u):
        t = tower_map.get(u.tower_id)
        p = project_map.get(t.project_id) if t else None
        return {
            "id": str(u.id), "unit_number": u.unit_number,
            "unit_type": u.unit_type, "bedrooms": u.bedrooms,
            "floor_number": u.floor_number, "facing": u.facing,
            "area_sqft": str(u.area_sqft) if u.area_sqft else None,
            "base_price": str(u.base_price) if u.base_price else None,
            "token_amount": str(u.token_amount) if u.token_amount else "20000",
            "emi_estimate": str(u.emi_estimate) if u.emi_estimate else None,
            "status": u.status,
            "is_trending": bool(u.is_trending) if u.is_trending is not None else False,
            "is_featured": bool(u.is_featured) if u.is_featured is not None else False,
            "view_count": u.view_count or 0,
            "tower_id": str(u.tower_id) if u.tower_id else None,
            "tower_name": t.name if t else None,
            "project_id": str(t.project_id) if t and t.project_id else None,
            "project_name": p.name if p else None,
        }

    return {
        "total": total, "page": page, "page_size": page_size,
        "items": [_row(u) for u in units]
    }



@router.get("/units/all")
async def list_units_all(
    page: int = 1, page_size: int = 20,
    status: str = "", unit_type: str = "",
    tower_id: str = "", project_id: str = "",
    db: AsyncSession = Depends(get_db), admin=Depends(verify_admin_token)
):
    """Alias for list_units — /units/all maps to the same list endpoint."""
    return await list_units(page=page, page_size=page_size, status=status,
                            unit_type=unit_type, db=db, admin=admin)

@router.get("/units/{unit_id}")
async def get_unit_admin(
    unit_id: UUID, db: AsyncSession = Depends(get_db), admin=Depends(verify_admin_token)
):
    result = await db.execute(select(Unit).where(Unit.id == unit_id))
    unit = result.scalar_one_or_none()
    if not unit: raise HTTPException(404, "Unit not found")
    cols = [c.name for c in Unit.__table__.columns]
    return {c: str(getattr(unit, c)) if hasattr(getattr(unit, c, None), 'hex') else getattr(unit, c) 
            for c in cols if c != 'embedding'}

@router.patch("/units/{unit_id}")
async def update_unit(
    unit_id: UUID, data: dict,
    db: AsyncSession = Depends(get_db), admin=Depends(verify_admin_token)
):
    result = await db.execute(select(Unit).where(Unit.id == unit_id))
    unit = result.scalar_one_or_none()
    if not unit: raise HTTPException(404, "Unit not found")

    decimal_fields = {"base_price", "emi_estimate", "down_payment", "price_per_sqft", "area_sqft", "carpet_area", "token_amount"}
    json_fields = {"dimensions", "images", "floor_plans", "amenities"}
    allowed = {
        "status", "base_price", "emi_estimate", "down_payment", "price_per_sqft",
        "is_trending", "is_featured", "facing", "floor_number", "unit_type",
        "bedrooms", "bathrooms", "area_sqft", "carpet_area",
        "dimensions", "images", "floor_plan_img", "floor_plans",
        "video_url", "walkthrough_url", "amenities", "token_amount",
    }

    scalar_sets = {}
    json_sets = {}

    for k, v in data.items():
        if k not in allowed:
            continue
        if k in decimal_fields and v is not None:
            scalar_sets[k] = Decimal(str(v))
        elif k in json_fields:
            json_sets[k] = v if isinstance(v, list) else []
        else:
            scalar_sets[k] = v

    # Apply scalar updates via raw SQL to guarantee persistence
    all_sql_sets = {}
    for k, v in scalar_sets.items():
        all_sql_sets[k] = v
    for k, v in json_sets.items():
        all_sql_sets[k] = json.dumps(v) if isinstance(v, (list, dict)) else v

    if all_sql_sets:
        set_parts = ", ".join(f"{k} = :{k}" for k in all_sql_sets)
        params = {**all_sql_sets, "unit_id": str(unit_id)}
        await db.execute(
            text(f"UPDATE units SET {set_parts} WHERE id = :unit_id"),
            params
        )

    await db.commit()
    await db.refresh(unit)
    from decimal import Decimal as _D
    import uuid as _uuid
    def _sv(v):
        if isinstance(v, _D): return str(v)
        if isinstance(v, _uuid.UUID): return str(v)
        return v
    cols = [col.name for col in unit.__table__.columns if col.name != 'embedding']
    return {col: _sv(getattr(unit, col)) for col in cols}


@router.post("/units/bulk-dimensions", status_code=200)
async def bulk_upload_dimensions(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    admin=Depends(verify_admin_token),
):
    """
    Bulk upload room dimensions for multiple units via CSV or XLSX.
    Required columns: project_name, tower_name, unit_number, room, width, length
    Optional column:  unit (ft/m/in — defaults to ft)
    The combination of project_name + tower_name + unit_number is the unique key.
    """
    filename = (file.filename or "").lower()
    if not (filename.endswith(".csv") or filename.endswith(".xlsx")):
        raise HTTPException(400, "Only .csv or .xlsx files are accepted")

    content = await file.read()
    rows: list[dict] = []

    if filename.endswith(".xlsx"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
        except Exception as e:
            raise HTTPException(400, f"Failed to parse Excel file: {e}")
    else:
        text_content = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text_content))
        rows = [{k: (v.strip() if v else "") for k, v in row.items()} for row in reader]

    if not rows:
        raise HTTPException(400, "File is empty")

    # Validate required columns
    required = {"project_name", "tower_name", "unit_number", "room", "width", "length"}
    missing = required - set(rows[0].keys())
    if missing:
        raise HTTPException(400, f"Missing columns: {', '.join(sorted(missing))}")

    # Group dimension rows by (project_name, tower_name, unit_number)
    # Key: (project_name_lower, tower_name_lower, unit_number) → list of dim dicts
    grouped: dict[tuple, dict] = {}
    row_errors: list[str] = []

    for i, row in enumerate(rows, start=2):
        project_name = row.get("project_name", "").strip()
        tower_name   = row.get("tower_name", "").strip()
        unit_num     = row.get("unit_number", "").strip()
        room         = row.get("room", "").strip()

        if not project_name or not tower_name or not unit_num:
            row_errors.append(f"Row {i}: project_name, tower_name and unit_number are all required")
            continue
        if not room:
            row_errors.append(f"Row {i}: room is required")
            continue

        try:
            width  = float(row.get("width") or 0)
            length = float(row.get("length") or 0)
        except ValueError:
            row_errors.append(f"Row {i}: width and length must be numbers")
            continue

        dim_unit = row.get("unit", "ft").strip() or "ft"
        key = (project_name.lower(), tower_name.lower(), unit_num)
        if key not in grouped:
            grouped[key] = {
                "project_name": project_name,
                "tower_name":   tower_name,
                "unit_number":  unit_num,
                "dims": [],
            }
        grouped[key]["dims"].append({
            "room": room, "width": width, "length": length, "unit": dim_unit,
        })

    if not grouped:
        raise HTTPException(400, f"No valid rows found. Errors: {'; '.join(row_errors[:5])}")

    # Resolve projects (cache by name)
    project_cache: dict[str, any] = {}
    for key, entry in grouped.items():
        pname = key[0]
        if pname not in project_cache:
            r = await db.execute(
                select(Project).where(func.lower(Project.name) == pname)
            )
            project_cache[pname] = r.scalar_one_or_none()

    # Resolve towers scoped to their project (cache by project_id + tower_name)
    tower_cache: dict[tuple, any] = {}
    for key, entry in grouped.items():
        pname, tname, _ = key
        project = project_cache.get(pname)
        if not project:
            continue
        tcache_key = (str(project.id), tname)
        if tcache_key not in tower_cache:
            r = await db.execute(
                select(Tower).where(
                    Tower.project_id == project.id,
                    func.lower(Tower.name) == tname,
                )
            )
            tower_cache[tcache_key] = r.scalar_one_or_none()

    # Resolve units scoped to their tower
    unit_cache: dict[tuple, any] = {}
    for key, entry in grouped.items():
        pname, tname, unit_num = key
        project = project_cache.get(pname)
        if not project:
            continue
        tower = tower_cache.get((str(project.id), tname))
        if not tower:
            continue
        r = await db.execute(
            select(Unit).where(
                Unit.tower_id == tower.id,
                Unit.unit_number == unit_num,
            )
        )
        unit_cache[key] = r.scalar_one_or_none()

    # Apply updates
    updated = 0
    not_found: list[str] = []

    for key, entry in grouped.items():
        pname, tname, unit_num = key
        label = f"{entry['project_name']} / {entry['tower_name']} / {unit_num}"

        project = project_cache.get(pname)
        if not project:
            not_found.append(f"{label} (project not found)")
            continue

        tower = tower_cache.get((str(project.id), tname))
        if not tower:
            not_found.append(f"{label} (tower not found)")
            continue

        unit_obj = unit_cache.get(key)
        if not unit_obj:
            not_found.append(f"{label} (unit not found)")
            continue

        await db.execute(
            text("UPDATE units SET dimensions = :dims WHERE id = :uid"),
            {"dims": json.dumps(entry["dims"]), "uid": str(unit_obj.id)},
        )
        updated += 1

    await db.commit()
    return {
        "updated": updated,
        "not_found": not_found,
        "row_errors": row_errors,
        "total_rows": sum(len(e["dims"]) for e in grouped.values()),
    }


@router.get("/units/bulk-dimensions/template")
async def download_dimensions_template(admin=Depends(verify_admin_token)):
    """Return a sample CSV template for bulk dimension upload."""
    lines = [
        "project_name,tower_name,unit_number,room,width,length,unit",
        "Janapriya Heights,Tower A,A101,Master Bedroom,12.6,14.0,ft",
        "Janapriya Heights,Tower A,A101,Living Room,16.0,20.3,ft",
        "Janapriya Heights,Tower A,A101,Kitchen,10.0,12.0,ft",
        "Janapriya Heights,Tower A,A101,Balcony,6.0,10.0,ft",
        "Janapriya Heights,Tower A,A102,Master Bedroom,12.6,14.0,ft",
        "Janapriya Heights,Tower A,A102,Living Room,16.0,20.3,ft",
        "Janapriya Heights,Tower B,A101,Master Bedroom,11.6,13.0,ft",
    ]
    from fastapi.responses import Response
    return Response(
        content="\n".join(lines),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=dimensions_template.csv"},
    )


@router.get("/projects")
async def list_projects(
    db: AsyncSession = Depends(get_db), admin=Depends(verify_admin_token)
):
    result = await db.execute(select(Project).order_by(Project.created_at.desc()))
    projects = result.scalars().all()
    return {
        "total": len(projects),
        "items": [
            {
                "id": str(p.id), "name": p.name, "location": p.location,
                "city": p.city, "rera_number": p.rera_number,
                "is_active": p.is_active, "is_featured": p.is_featured,
                "created_at": p.created_at.isoformat() if p.created_at else None,
            }
            for p in projects
        ]
    }


from backend.app.models.tower import Tower

@router.get("/towers")
async def list_towers(
    project_id: str = "",
    db: AsyncSession = Depends(get_db), admin=Depends(verify_admin_token)
):
    from sqlalchemy import select as sel
    from backend.app.models.project import Project
    q = sel(Tower)
    if project_id:
        from uuid import UUID as _UUID
        q = q.where(Tower.project_id == _UUID(project_id))
    result = await db.execute(q.order_by(Tower.created_at.desc()))
    towers = result.scalars().all()
    return {
        "total": len(towers),
        "items": [
            {
                "id": str(t.id), "name": t.name,
                "project_id": str(t.project_id),
                "total_floors": t.total_floors,
                "total_units": t.total_units,
                "amenities": t.amenities or [],
                "brochure_url": t.brochure_url,
                "is_active": bool(t.is_active) if t.is_active is not None else True,
            }
            for t in towers
        ]
    }

@router.get("/towers/list")
async def list_towers_alias(
    project_id: str = "",
    db: AsyncSession = Depends(get_db), admin=Depends(verify_admin_token)
):
    return await list_towers(project_id=project_id, db=db, admin=admin)

@router.get("/towers/{tower_id}")
async def get_tower(
    tower_id: UUID, db: AsyncSession = Depends(get_db), admin=Depends(verify_admin_token)
):
    result = await db.execute(select(Tower).where(Tower.id == tower_id))
    tower = result.scalar_one_or_none()
    if not tower: raise HTTPException(404, "Tower not found")
    return {
        "id": str(tower.id), "name": tower.name,
        "project_id": str(tower.project_id),
        "total_floors": tower.total_floors,
        "total_units": tower.total_units,
        "amenities": tower.amenities or [],
        "video_url": tower.video_url,
        "walkthrough_url": tower.walkthrough_url,
        "brochure_url": tower.brochure_url,
    }

@router.patch("/towers/{tower_id}")
async def update_tower(
    tower_id: UUID, data: dict,
    db: AsyncSession = Depends(get_db), admin=Depends(verify_admin_token)
):
    result = await db.execute(select(Tower).where(Tower.id == tower_id))
    tower = result.scalar_one_or_none()
    if not tower: raise HTTPException(404, "Tower not found")
    scalar_fields = {"name", "description", "video_url", "walkthrough_url", "svg_floor_plan"}
    int_fields = {"total_floors", "total_units"}
    bool_fields = {"is_active"}
    json_fields = {"amenities", "images", "floor_plans"}
    allowed = scalar_fields | int_fields | bool_fields | json_fields | {"brochure_url"}
    import json as _json
    json_updates = {}
    scalar_updates = {}
    for k, v in data.items():
        if k not in allowed:
            continue
        if k in json_fields:
            json_updates[k] = v if isinstance(v, list) else []
        elif k in int_fields:
            try:
                scalar_updates[k] = int(v) if v not in (None, "", []) else None
            except (ValueError, TypeError):
                pass
        elif k in bool_fields:
            # Coerce string "true"/"false" to bool
            if isinstance(v, bool):
                scalar_updates[k] = v
            elif isinstance(v, str):
                scalar_updates[k] = v.lower() in ("true", "1", "yes")
            elif isinstance(v, int):
                scalar_updates[k] = bool(v)
        else:
            scalar_updates[k] = v
    # Apply scalar updates via setattr
    for k, v in scalar_updates.items():
        if v is not None:
            setattr(tower, k, v)
    # Apply JSON updates via raw SQL
    if json_updates:
        set_parts = ", ".join(f"{k} = :{k}" for k in json_updates)
        params = {k: _json.dumps(v) for k, v in json_updates.items()}
        params["tower_id"] = str(tower_id)
        await db.execute(text(f"UPDATE towers SET {set_parts} WHERE id = :tower_id"), params)
    await db.commit()
    await db.refresh(tower)
    return {
        "id": str(tower.id), "name": tower.name,
        "total_floors": tower.total_floors, "total_units": tower.total_units,
        "description": tower.description,
        "is_active": bool(tower.is_active),
        "amenities": tower.amenities or [], "brochure_url": tower.brochure_url,
        "video_url": tower.video_url, "walkthrough_url": tower.walkthrough_url,
    }


@router.get("/projects/list")
async def list_projects_alias(
    db: AsyncSession = Depends(get_db), admin=Depends(verify_admin_token)
):
    return await list_projects(db=db, admin=admin)

@router.get("/projects/{project_id}")
async def get_project_admin(
    project_id: UUID, db: AsyncSession = Depends(get_db), admin=Depends(verify_admin_token)
):
    from backend.app.models.project import Project as Proj
    result = await db.execute(select(Proj).where(Proj.id == project_id))
    proj = result.scalar_one_or_none()
    if not proj: raise HTTPException(404, "Project not found")
    import json as _json
    from sqlalchemy import text as _t
    row = (await db.execute(_t("SELECT * FROM projects WHERE id=:id"), {"id": str(project_id)})).mappings().fetchone()
    return dict(row) if row else {}


# ── Test Notification Endpoints ──

class TestNotificationRequest(BaseModel):
    channel: str  # "email", "sms", "whatsapp"
    phone: Optional[str] = None
    email: Optional[str] = None
    name: str = "Test User"
    unit_number: str = "A2-TEST"
    booking_id: str = "TEST1234"
    amount: str = "20000"
    transaction_id: str = "pay_test_123456"


@router.post("/test-notification")
async def test_notification(
    data: TestNotificationRequest,
    admin=Depends(verify_admin_token),
):
    """Send a test booking notification via the specified channel."""
    results = {}

    if data.channel == "email":
        if not data.email:
            raise HTTPException(400, "Email address is required")
        from backend.app.services.email import send_booking_confirmation_email
        ok = await send_booking_confirmation_email(
            to_email=data.email,
            customer_name=data.name,
            unit_number=data.unit_number,
            project_name="Janapriya Upscale",
            booking_amount=f"Rs.{data.amount}",
            total_amount="Rs.30,00,000",
            payment_id=data.transaction_id,
            booking_id=data.booking_id,
            booked_at="04 Apr 2026",
        )
        results["email"] = "sent" if ok else "failed"

    elif data.channel == "sms":
        if not data.phone:
            raise HTTPException(400, "Phone number is required")
        from backend.app.services.sms import send_booking_confirmation_sms, send_payment_confirmation_sms
        ok1 = await send_booking_confirmation_sms(
            phone=data.phone, unit_number=data.unit_number, booking_id=data.booking_id,
        )
        ok2 = await send_payment_confirmation_sms(
            phone=data.phone, amount=data.amount, transaction_id=data.transaction_id,
        )
        results["sms_booking"] = "sent" if ok1 else "failed"
        results["sms_payment"] = "sent" if ok2 else "failed"

    elif data.channel == "whatsapp":
        if not data.phone:
            raise HTTPException(400, "Phone number is required")
        from backend.app.services.whatsapp import send_booking_confirmation_whatsapp, send_payment_confirmation_whatsapp
        ok1 = await send_booking_confirmation_whatsapp(
            phone=data.phone, customer_name=data.name,
            unit_number=data.unit_number, booking_id=data.booking_id,
        )
        ok2 = await send_payment_confirmation_whatsapp(
            phone=data.phone, customer_name=data.name,
            amount=data.amount, transaction_id=data.transaction_id,
        )
        results["whatsapp_booking"] = "sent" if ok1 else "failed"
        results["whatsapp_payment"] = "sent" if ok2 else "failed"
    else:
        raise HTTPException(400, "Channel must be 'email', 'sms', or 'whatsapp'")

    return {"channel": data.channel, "results": results}

